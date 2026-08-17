import openpyxl
from openpyxl.styles import PatternFill, Font, Alignment
from collections import Counter, defaultdict
import os

WX_FILE = r"C:\Users\Administrator\Desktop\单次核对\微信流水.xlsx"
RT_FILE = r"C:\Users\Administrator\Desktop\单次核对\零售单.xlsx"
OUT_FILE = r"C:\Users\Administrator\Desktop\单次核对\微信流水_零售单_核对报告.xlsx"

# ============ 1. Read WeChat flow ============
wb_wx = openpyxl.load_workbook(WX_FILE, read_only=True, data_only=True)
ws_wx = wb_wx["Sheet1"]
wx_header = [c.value for c in next(ws_wx.iter_rows(min_row=1, max_row=1))]
idx_wx = {h: i for i, h in enumerate(wx_header)}

wx_records = []
for i, row in enumerate(ws_wx.iter_rows(min_row=2, values_only=True)):
    if row is None or all(v is None for v in row):
        continue
    wx_records.append({
        "row": i + 2,
        "交易时间": row[idx_wx["交易时间"]],
        "门店": row[idx_wx["门店"]],
        "交易类型": row[idx_wx["交易类型"]],
        "交易状态": row[idx_wx["交易状态"]],
        "订单金额": row[idx_wx["订单金额"]],
        "应结订单金额": row[idx_wx["应结订单金额"]],
        "退款金额": row[idx_wx["退款金额"]],
        "微信订单号": row[idx_wx["微信订单号"]],
        "商户订单号": row[idx_wx["商户订单号"]],
        "商品名称": row[idx_wx["商品名称"]],
        "费率备注": row[idx_wx["费率备注"]],
    })
wb_wx.close()

# Categorize WeChat records:
# - 实物支付: 商品名称='曹宅店' AND 费率备注='实物' AND 交易状态='SUCCESS'
# - 充值: 费率备注='充值'
# - 退款: 交易状态='REFUND' or 退款金额>0
wx_pay_sales = [r for r in wx_records if r["费率备注"] == "实物" and r["交易状态"] == "SUCCESS"]
wx_recharge = [r for r in wx_records if r["费率备注"] == "充值"]
wx_refund = [r for r in wx_records if r["交易状态"] == "REFUND" or (r["退款金额"] not in (None, 0, 0.0) and r["退款金额"])]
print(f"微信流水分类:")
print(f"  实物销售支付(SUCCESS): {len(wx_pay_sales)} 笔, 金额合计: {sum(float(r['订单金额']) for r in wx_pay_sales):.2f}")
print(f"  充值交易: {len(wx_recharge)} 笔, 金额合计: {sum(float(r['订单金额']) for r in wx_recharge):.2f}")
print(f"  退款记录: {len(wx_refund)} 笔")

# ============ 2. Read Retail REPORT0 ============
wb_rt = openpyxl.load_workbook(RT_FILE, read_only=True, data_only=True)
ws_rt = wb_rt["REPORT0"]
rt_header = [c.value for c in next(ws_rt.iter_rows(min_row=1, max_row=1))]
idx_rt = {h: i for i, h in enumerate(rt_header)}

rt_records = []
for i, row in enumerate(ws_rt.iter_rows(min_row=2, values_only=True)):
    if row is None or all(v is None for v in row):
        continue
    # skip footer/合计 row (门店名称 is None or '合计')
    if row[idx_rt["门店名称"]] is None or row[idx_rt["门店名称"]] == "合计":
        continue
    rt_records.append({
        "row": i + 2,
        "门店名称": row[idx_rt["门店名称"]],
        "营业日期": row[idx_rt["营业日期"]],
        "销售单号": row[idx_rt["销售单号"]],
        "渠道": row[idx_rt["渠道"]],
        "实收金额": row[idx_rt["实收金额"]],
        "微信": row[idx_rt["微信"]],
        "合计": row[idx_rt["合计"]],
        "单据类型": row[idx_rt["单据类型"]],
        "系统日期": row[idx_rt["系统日期"]],
        "系统时间": row[idx_rt["系统时间"]],
    })
wb_rt.close()

# Retail records with WeChat amount > 0 (both 0 销售 and 3 订单)
rt_wx = [r for r in rt_records if r["微信"] not in (None, 0, 0.0) and float(r["微信"] or 0) > 0]
rt_sales_wx = [r for r in rt_wx if r["单据类型"] and "销售" in str(r["单据类型"])]
rt_order_wx = [r for r in rt_wx if r["单据类型"] and "订单" in str(r["单据类型"])]
print(f"\n零售单REPORT0分类(已剔除合计行):")
print(f"  总记录数: {len(rt_records)}")
print(f"  有微信金额的记录: {len(rt_wx)} 笔, 微信金额合计: {sum(float(r['微信']) for r in rt_wx):.2f}")
print(f"    其中 '0 销售': {len(rt_sales_wx)} 笔, 合计: {sum(float(r['微信']) for r in rt_sales_wx):.2f}")
print(f"    其中 '3 订单': {len(rt_order_wx)} 笔, 合计: {sum(float(r['微信']) for r in rt_order_wx):.2f}")

# ============ 3. Amount-based matching (实物支付 vs 全部零售单微信金额) ============
def norm_amt(v):
    if v is None:
        return None
    try:
        return round(float(v), 2)
    except Exception:
        return None

# Match WeChat 实物支付 (94 records) against retail WeChat records (93 records: 89 销售 + 4 订单)
wx_amt_counter = Counter()
for r in wx_pay_sales:
    a = norm_amt(r["订单金额"])
    if a is not None:
        wx_amt_counter[a] += 1

rt_amt_counter = Counter()
for r in rt_wx:
    a = norm_amt(r["微信"])
    if a is not None:
        rt_amt_counter[a] += 1

all_amts = set(wx_amt_counter) | set(rt_amt_counter)
matched_count = 0
detail_rows = []
for a in sorted(all_amts):
    w = wx_amt_counter.get(a, 0)
    r = rt_amt_counter.get(a, 0)
    m = min(w, r)
    matched_count += m
    detail_rows.append((a, w, r, m, w - m, r - m))

print(f"\n=== 金额匹配结果（微信实物支付 vs 零售单微信金额）===")
print(f"微信实物支付: {sum(wx_amt_counter.values())} 笔, 金额: {sum(a*c for a,c in wx_amt_counter.items()):.2f}")
print(f"零售单微信金额: {sum(rt_amt_counter.values())} 笔, 金额: {sum(a*c for a,c in rt_amt_counter.items()):.2f}")
print(f"按金额匹配上: {matched_count} 笔")
print(f"微信侧未匹配: {sum(wl for *_, wl, rl in detail_rows)} 笔, 金额: {sum(wl*a for a,w,r,m,wl,rl in detail_rows):.2f}")
print(f"零售单侧未匹配: {sum(rl for *_, wl, rl in detail_rows)} 笔, 金额: {sum(rl*a for a,w,r,m,wl,rl in detail_rows):.2f}")

print(f"\n金额不匹配明细:")
print(f"{'金额':>10} | {'微信笔数':>6} | {'零售单笔数':>8} | {'匹配':>4} | {'微信剩':>6} | {'零售单剩':>8}")
for a, w, r, m, wl, rl in detail_rows:
    if wl != 0 or rl != 0:
        print(f"{a:>10.2f} | {w:>6} | {r:>8} | {m:>4} | {wl:>6} | {rl:>8}")

# ============ 4. Per-record pairing ============
wx_by_amt = defaultdict(list)
for r in wx_pay_sales:
    a = norm_amt(r["订单金额"])
    if a is not None:
        wx_by_amt[a].append(r)

rt_by_amt = defaultdict(list)
for r in rt_wx:
    a = norm_amt(r["微信"])
    if a is not None:
        rt_by_amt[a].append(r)

matched_pairs = []
wx_leftover = []
rt_leftover = []
for a in sorted(all_amts):
    wxs = list(wx_by_amt.get(a, []))
    rts = list(rt_by_amt.get(a, []))
    n = min(len(wxs), len(rts))
    for i in range(n):
        matched_pairs.append((wxs[i], rts[i], a))
    if len(wxs) > n:
        wx_leftover.extend(wxs[n:])
    if len(rts) > n:
        rt_leftover.extend(rts[n:])

print(f"\n笔级配对: 成功 {len(matched_pairs)} 笔, 微信剩 {len(wx_leftover)} 笔, 零售单剩 {len(rt_leftover)} 笔")
print(f"\n微信侧未匹配记录:")
for r in wx_leftover:
    print(f"  金额={r['订单金额']}, 时间={r['交易时间']}, 商品={r['商品名称']}, 备注={r['费率备注']}")
print(f"\n零售单侧未匹配记录:")
for r in rt_leftover:
    print(f"  金额={r['微信']}, 单号={r['销售单号']}, 类型={r['单据类型']}")

# ============ 5. Write output report ============
wb_out = openpyxl.Workbook()
header_fill = PatternFill("solid", fgColor="305496")
header_font = Font(color="FFFFFF", bold=True)
center = Alignment(horizontal="center", vertical="center", wrap_text=True)

# Sheet 1: 核对汇总
ws_sum = wb_out.active
ws_sum.title = "核对汇总"
sum_data = [
    ["项目", "数值", "说明"],
    ["微信流水总记录数", len(wx_records), "全部记录"],
    ["微信-实物销售支付(SUCCESS)", len(wx_pay_sales), "商品名称=曹宅店, 费率备注=实物, 交易状态=SUCCESS"],
    ["微信-充值交易", len(wx_recharge), "费率备注=充值, 不参与销售单匹配"],
    ["微信-退款记录", len(wx_refund), "交易状态=REFUND 或 退款金额>0"],
    ["", "", ""],
    ["零售单REPORT0总记录数", len(rt_records), "已剔除合计行"],
    ["零售单-有微信金额的记录", len(rt_wx), "微信列>0"],
    ["  其中 '0 销售'", len(rt_sales_wx), ""],
    ["  其中 '3 订单'", len(rt_order_wx), ""],
    ["", "", ""],
    ["【金额匹配结果】微信实物支付 vs 零售单微信金额", "", ""],
    ["微信侧总笔数", sum(wx_amt_counter.values()), ""],
    ["微信侧总金额", round(sum(a*c for a,c in wx_amt_counter.items()), 2), ""],
    ["零售单侧总笔数", sum(rt_amt_counter.values()), ""],
    ["零售单侧总金额", round(sum(a*c for a,c in rt_amt_counter.items()), 2), ""],
    ["匹配成功笔数", matched_count, ""],
    ["微信侧未匹配笔数", sum(wl for *_, wl, rl in detail_rows), ""],
    ["微信侧未匹配金额", round(sum(wl*a for a,w,r,m,wl,rl in detail_rows), 2), ""],
    ["零售单侧未匹配笔数", sum(rl for *_, wl, rl in detail_rows), ""],
    ["零售单侧未匹配金额", round(sum(rl*a for a,w,r,m,wl,rl in detail_rows), 2), ""],
    ["两侧金额合计差异(微信-零售单)", round(sum(a*c for a,c in wx_amt_counter.items()) - sum(a*c for a,c in rt_amt_counter.items()), 2), ""],
    ["", "", ""],
    ["【充值类说明】", "", "充值交易不会出现在零售销售单中,属正常差异"],
    ["充值笔数", len(wx_recharge), ""],
    ["充值金额合计", round(sum(float(r['订单金额']) for r in wx_recharge), 2), ""],
]
for r in sum_data:
    ws_sum.append(r)
for c in ws_sum[1]:
    c.fill = header_fill
    c.font = header_font
    c.alignment = center
ws_sum.column_dimensions["A"].width = 42
ws_sum.column_dimensions["B"].width = 14
ws_sum.column_dimensions["C"].width = 50

# Sheet 2: 金额维度对比
ws_amt = wb_out.create_sheet("金额维度对比")
ws_amt.append(["金额", "微信笔数", "零售单笔数", "匹配笔数", "微信剩余笔数", "零售单剩余笔数", "微信剩余金额", "零售单剩余金额", "差异标记"])
for a, w, r, m, wl, rl in detail_rows:
    flag = ""
    if wl > 0 and rl > 0:
        flag = "两侧都有剩余"
    elif wl > 0:
        flag = "仅微信有"
    elif rl > 0:
        flag = "仅零售单有"
    ws_amt.append([a, w, r, m, wl, rl, round(wl*a, 2), round(rl*a, 2), flag])
for c in ws_amt[1]:
    c.fill = header_fill
    c.font = header_font
    c.alignment = center
fill_wx = PatternFill("solid", fgColor="FFE699")
fill_rt = PatternFill("solid", fgColor="C6E0B4")
fill_both = PatternFill("solid", fgColor="F8CBAD")
for row in ws_amt.iter_rows(min_row=2):
    flag = row[8].value
    if flag == "仅微信有":
        for c in row: c.fill = fill_wx
    elif flag == "仅零售单有":
        for c in row: c.fill = fill_rt
    elif flag == "两侧都有剩余":
        for c in row: c.fill = fill_both
for i, w in enumerate([10, 10, 12, 10, 14, 16, 14, 16, 16], 1):
    ws_amt.column_dimensions[openpyxl.utils.get_column_letter(i)].width = w

# Sheet 3: 配对明细
ws_pair = wb_out.create_sheet("配对明细")
ws_pair.append([
    "配对序号", "金额",
    "微信_交易时间", "微信_门店", "微信_微信订单号", "微信_商户订单号", "微信_商品名称", "微信_费率备注",
    "零售单_门店", "零售单_营业日期", "零售单_销售单号", "零售单_渠道", "零售单_单据类型", "零售单_实收金额", "零售单_微信金额", "零售单_系统时间"
])
for i, (wx, rt, a) in enumerate(matched_pairs, 1):
    ws_pair.append([
        i, a,
        str(wx["交易时间"]), wx["门店"], str(wx["微信订单号"]), str(wx["商户订单号"]), wx["商品名称"], wx["费率备注"],
        rt["门店名称"], rt["营业日期"], rt["销售单号"], rt["渠道"], rt["单据类型"], rt["实收金额"], rt["微信"], rt["系统时间"],
    ])
for c in ws_pair[1]:
    c.fill = header_fill
    c.font = header_font
    c.alignment = center
for i, w in enumerate([8, 10, 20, 10, 28, 28, 12, 12, 10, 12, 24, 16, 12, 10, 10, 10], 1):
    ws_pair.column_dimensions[openpyxl.utils.get_column_letter(i)].width = w

# Sheet 4: 微信侧未匹配
ws_wxl = wb_out.create_sheet("微信侧未匹配")
ws_wxl.append(["序号", "金额", "交易时间", "门店", "微信订单号", "商户订单号", "交易类型", "交易状态", "商品名称", "费率备注"])
for i, wx in enumerate(wx_leftover, 1):
    ws_wxl.append([i, norm_amt(wx["订单金额"]), str(wx["交易时间"]), wx["门店"], str(wx["微信订单号"]), str(wx["商户订单号"]),
                   wx["交易类型"], wx["交易状态"], wx["商品名称"], wx["费率备注"]])
for c in ws_wxl[1]:
    c.fill = header_fill
    c.font = header_font
    c.alignment = center
for i, w in enumerate([6, 10, 20, 10, 28, 28, 12, 12, 14, 12], 1):
    ws_wxl.column_dimensions[openpyxl.utils.get_column_letter(i)].width = w

# Sheet 5: 零售单侧未匹配
ws_rtl = wb_out.create_sheet("零售单侧未匹配")
ws_rtl.append(["序号", "金额", "门店", "营业日期", "销售单号", "渠道", "实收金额", "单据类型", "系统日期", "系统时间"])
for i, rt in enumerate(rt_leftover, 1):
    ws_rtl.append([i, norm_amt(rt["微信"]), rt["门店名称"], rt["营业日期"], rt["销售单号"], rt["渠道"],
                   rt["实收金额"], rt["单据类型"], rt["系统日期"], rt["系统时间"]])
for c in ws_rtl[1]:
    c.fill = header_fill
    c.font = header_font
    c.alignment = center
for i, w in enumerate([6, 10, 10, 12, 24, 16, 10, 14, 12, 10], 1):
    ws_rtl.column_dimensions[openpyxl.utils.get_column_letter(i)].width = w

# Sheet 6: 微信充值交易
ws_rc = wb_out.create_sheet("微信充值交易")
ws_rc.append(["序号", "交易时间", "门店", "微信订单号", "商户订单号", "订单金额", "商品名称", "费率备注"])
for i, wx in enumerate(wx_recharge, 1):
    ws_rc.append([i, str(wx["交易时间"]), wx["门店"], str(wx["微信订单号"]), str(wx["商户订单号"]),
                  wx["订单金额"], wx["商品名称"], wx["费率备注"]])
for c in ws_rc[1]:
    c.fill = header_fill
    c.font = header_font
    c.alignment = center
for i, w in enumerate([6, 20, 10, 28, 28, 10, 14, 12], 1):
    ws_rc.column_dimensions[openpyxl.utils.get_column_letter(i)].width = w

# Sheet 7: 微信退款记录
ws_rf = wb_out.create_sheet("微信退款记录")
ws_rf.append(["序号", "交易时间", "门店", "微信订单号", "商户订单号", "交易类型", "交易状态", "订单金额", "退款金额", "商品名称", "费率备注"])
for i, wx in enumerate(wx_refund, 1):
    ws_rf.append([i, str(wx["交易时间"]), wx["门店"], str(wx["微信订单号"]), str(wx["商户订单号"]),
                  wx["交易类型"], wx["交易状态"], wx["订单金额"], wx["退款金额"], wx["商品名称"], wx["费率备注"]])
for c in ws_rf[1]:
    c.fill = header_fill
    c.font = header_font
    c.alignment = center
for i, w in enumerate([6, 20, 10, 28, 28, 12, 12, 10, 10, 14, 12], 1):
    ws_rf.column_dimensions[openpyxl.utils.get_column_letter(i)].width = w

wb_out.save(OUT_FILE)
print(f"\n报告已保存: {OUT_FILE}")
