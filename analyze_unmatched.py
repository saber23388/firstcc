import openpyxl
from collections import Counter

WX_FILE = r"C:\Users\Administrator\Desktop\单次核对\微信流水.xlsx"
RT_FILE = r"C:\Users\Administrator\Desktop\单次核对\零售单.xlsx"

# Read WeChat
wb = openpyxl.load_workbook(WX_FILE, read_only=True, data_only=True)
ws = wb["Sheet1"]
hdr = [c.value for c in next(ws.iter_rows(min_row=1, max_row=1))]
idx = {h: i for i, h in enumerate(hdr)}
wx_recs = []
for row in ws.iter_rows(min_row=2, values_only=True):
    if row is None or all(v is None for v in row):
        continue
    wx_recs.append(row)
wb.close()

print("=== 微信流水 商品名称 / 费率备注 分布 ===")
goods_stat = Counter(row[idx["商品名称"]] for row in wx_recs)
note_stat = Counter(row[idx["费率备注"]] for row in wx_recs)
print(f"商品名称分布: {dict(goods_stat)}")
print(f"费率备注分布: {dict(note_stat)}")

print("\n=== 微信流水中金额 >= 100 的记录（疑似充值/大额）===")
print(f"{'交易时间':<22} {'门店':<8} {'订单金额':>10} {'商品名称':<12} {'费率备注':<10} {'交易状态':<10}")
for row in wx_recs:
    amt = row[idx["订单金额"]]
    if amt is not None and float(amt) >= 100:
        print(f"{str(row[idx['交易时间']]):<22} {str(row[idx['门店']]):<8} {amt:>10} {str(row[idx['商品名称']]):<12} {str(row[idx['费率备注']]):<10} {str(row[idx['交易状态']]):<10}")

# Read Retail REPORT0 - check "3 订单" records
wb2 = openpyxl.load_workbook(RT_FILE, read_only=True, data_only=True)
ws2 = wb2["REPORT0"]
hdr2 = [c.value for c in next(ws2.iter_rows(min_row=1, max_row=1))]
idx2 = {h: i for i, h in enumerate(hdr2)}
rt_recs = []
for row in ws2.iter_rows(min_row=2, values_only=True):
    if row is None or all(v is None for v in row):
        continue
    rt_recs.append(row)
wb2.close()

print("\n=== 零售单 REPORT0 中 '3 订单' 类型的记录 ===")
print(f"{'门店':<8} {'营业日期':<10} {'销售单号':<24} {'渠道':<16} {'实收金额':>10} {'微信':>10} {'单据类型':<10}")
for row in rt_recs:
    bt = row[idx2["单据类型"]]
    if bt and "订单" in str(bt):
        print(f"{str(row[idx2['门店名称']]):<8} {str(row[idx2['营业日期']]):<10} {str(row[idx2['销售单号']]):<24} {str(row[idx2['渠道']]):<16} {str(row[idx2['实收金额']]):>10} {str(row[idx2['微信']]):>10} {str(bt):<10}")

print("\n=== 零售单 REPORT0 中 单据类型为 None 的记录 ===")
for row in rt_recs:
    bt = row[idx2["单据类型"]]
    if bt is None:
        print(row[:19])

print("\n=== 零售单 '3 订单' 记录的微信金额合计 ===")
total_wx_order = 0
cnt = 0
for row in rt_recs:
    bt = row[idx2["单据类型"]]
    if bt and "订单" in str(bt):
        v = row[idx2["微信"]]
        if v not in (None, 0, 0.0):
            total_wx_order += float(v)
            cnt += 1
print(f"3订单记录中含微信金额的笔数: {cnt}, 合计: {total_wx_order:.2f}")
