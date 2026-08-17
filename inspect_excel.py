import openpyxl
import os

files = [
    r"C:\Users\Administrator\Desktop\单次核对\微信流水.xlsx",
    r"C:\Users\Administrator\Desktop\单次核对\零售单.xlsx",
]

for f in files:
    print(f"\n{'='*70}")
    print(f"File: {f}")
    print(f"Exists: {os.path.exists(f)}")
    if not os.path.exists(f):
        continue
    wb = openpyxl.load_workbook(f, read_only=True, data_only=True)
    print(f"Sheet names: {wb.sheetnames}")
    for sh_name in wb.sheetnames:
        ws = wb[sh_name]
        print(f"\n--- Sheet: {sh_name} (max_row={ws.max_row}, max_col={ws.max_column}) ---")
        # print first 5 rows
        for i, row in enumerate(ws.iter_rows(values_only=True)):
            if i >= 8:
                break
            print(f"Row{i}: {row}")
    wb.close()
